import time

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text

from classes.DataBase import DataBase
from openpyxl import load_workbook


def get_schema_names():
    wb: Workbook = load_workbook("xlsx/Gosprogrammy_v_tablitse_7.xlsx")
    ws: Worksheet = wb.worksheets[0]
    rows = list(ws.rows)
    schema_names = dict()
    for row in rows:
        schema_names[row[2].value] = row[1].value

    return schema_names


class DataBaseTable7(DataBase):
    def __init__(self, min_row, max_row):
        super().__init__()
        print("Получение схем")
        self.schema_names: dict = get_schema_names()
        print("Схемы получены")
        print("Получение показателей по схемам")
        self.indicators_by_gp = self.get_indicators_by_gp(self.schema_names)
        print("Схемы по показателям получены\n")
        self.end_indicator_dict = dict()
        self.wb = load_workbook("xlsx/Tablitsa_7.xlsx")
        print("Таблица определена")
        self.ws = self.wb.worksheets[0]
        print("Страница определена")

        self.imported_rows = []
        self.editable_columns = [
            'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V'
        ]

        for i in list(range(8, 2638))[::10]:
            self.min_row = i
            self.rows = list(self.ws.iter_rows(min_row=i, max_row=i + 10))
            self.find_all_for_table_7()
            del self.rows
            time.sleep(0.001)
        print("Таблица xlsx загружена")

    def find_all_for_table_7(self):
        schema_gp_list = [
            (row_num, (row[2].value, row[9].value))
            for row_num, row in enumerate(self.rows, start=self.min_row)
            if row[2].value
        ]

        for sgp in schema_gp_list:
            if sgp[1][0] and sgp[1][1]:
                real_schema_name = self.schema_names[sgp[1][0].strip()]
                sgp_indicators = sgp[1][1].split('\n')
                for sgp_indicator in sgp_indicators:
                    for indicator_name in self.indicators_by_gp[real_schema_name]:
                        if indicator_name in sgp_indicator and indicator_name not in [" ", "\n"]:
                            self.import_info(sgp[0], self.get_indicators_by_name(
                                real_schema_name,
                                indicator_name
                            ))
                            break

    def import_info(self, row_num, info):
        unit_of_measurement = '-'
        for year_info in info:
            if year_info[-1] != "-":
                unit_of_measurement = year_info[-1][-1]
                break
        self.ws[f"K{row_num}"] = unit_of_measurement

        letters = iter(self.editable_columns[1:])
        for year_info in info:
            try:
                if year_info[-1] != "-":
                    if row_num not in self.imported_rows:
                        self.ws[f"{next(letters)}{row_num}"] = year_info[-1][0]
                        self.ws[f"{next(letters)}{row_num}"].value = year_info[-1][1]
                    else:
                        letter = next(letters)
                        value = str(self.ws[f"{letter}{row_num}"].value) + f"\n{year_info[-1][0]}"
                        self.ws[f"{letter}{row_num}"] = value

                        letter = next(letters)
                        value = str(self.ws[f"{letter}{row_num}"].value) + f"\n{year_info[-1][1]}"
                        self.ws[f"{letter}{row_num}"].value = value
                else:
                    if row_num not in self.imported_rows:
                        self.ws[f"{next(letters)}{row_num}"].value = "-"
                        self.ws[f"{next(letters)}{row_num}"].value = "-"
                    else:
                        letter = next(letters)
                        value = str(self.ws[f"{letter}{row_num}"].value) + "\n-"
                        self.ws[f"{letter}{row_num}"].value = value

                        letter = next(letters)
                        value = str(self.ws[f"{letter}{row_num}"].value) + "\n-"
                        self.ws[f"{letter}{row_num}"].value = value

            except StopIteration:
                self.imported_rows.append(row_num)
                print(row_num)

    def get_indicators_by_gp(self, schema_list):
        indicators_by_gp = dict()
        for key, schema in schema_list.items():
            indicators_by_gp[schema] = self.get_GP_indicator_names(schema)

        return indicators_by_gp

    def get_indicators_by_name(self, schema_name, ind_name: str):
        ind_name = ind_name.replace('%', '%%') if '%' in ind_name else ind_name

        tables_with_titles = self.get_table_list_by_schema_and_table_name(schema_name, "names_indicators____")
        tables_with_indicators = self.get_table_list_by_schema_and_table_name(schema_name, "indicators____")

        titles_id_by_years = [
            (table[-4:],
             self.conn.execute(text(
                 f"SELECT id, type FROM " +
                 f'"{schema_name}"' + f".{table} WHERE '{ind_name}' LIKE title_indication")
             ).first()
             ) for table in tables_with_titles
        ]

        indicators_by_years = []
        for year in range(len(titles_id_by_years)):
            if titles_id_by_years[year][1] is not None:
                indicators_by_years.append(
                    (titles_id_by_years[year][0],
                     self.conn.execute(text(
                         f"SELECT plan, fact, unit_of_measurement FROM " +
                         f'"{schema_name}"' +
                         f".{tables_with_indicators[year]} WHERE id = {titles_id_by_years[year][1][0]}")
                     ).first()
                     )
                )
            else:
                indicators_by_years.append((titles_id_by_years[year][0], "-"))

        return indicators_by_years


table_parsing = DataBaseTable7(8, 18)
table_parsing.wb.save("output.xlsx")
