import os

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Side, Border, PatternFill, NamedStyle, Alignment

"""
Для корректной работы надо сначала запустить метод CYR, потом GP. 
Полученный файл будет называться CYR_GP_2
"""


def create_diagram(sheet, min_row, max_row, min_col, max_col, row, col):
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Динамика изменения значений показателя"
    chart1.y_axis.delete = False
    chart1.x_axis.delete = False
    data = Reference(sheet, min_col=min_col + 1, max_col=max_col, min_row=min_row - 1, max_row=max_row - 1)
    categor = Reference(sheet, min_col=min_col, min_row=min_row, max_row=max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(categor)
    chart1.width = 14
    sheet.add_chart(chart1, 'ABCDEFG'[row - 1] + str(col))


class TableBlockConstructor:
    def __init__(self, first, second):
        self.first = first
        self.second = second

    def fill_CYR_GP_sample(self, indicator, event_id, event, main_event_id, main_event, subprogram_id, subprogram,
                           type, unit, years, plan_marks, fact_marks, year_comment, comment,
                           indicator_rf, task_id, task, cyr_id, cyr, indicator_vo, gosprogram, response_obj, path):
        path_test = self.fill_CYR_sample(2, './tableTemplates/CYR-GP.xlsx', indicator_rf, task_id, task, cyr_id, cyr,
                                         indicator_vo, gosprogram, response_obj, '../test.xlsx')
        self.fill_GP_sample(5, path_test, indicator, event_id, event, main_event_id, main_event, subprogram_id,
                            subprogram,
                            type, unit, years, plan_marks, fact_marks, year_comment, comment, path)
        os.remove(path_test)

    def fill_GP_GP_sample(self, indicator, event_id, event, main_event_id, main_event, subprogram_id, subprogram,
                          type, unit, years, plan_marks, fact_marks, year_comment, comment, indicator_1, event_id_1,
                          event_1,
                          main_event_id_1, main_event_1, subprogram_id_1, subprogram_1, type_1, unit_1, years_1,
                          plan_marks_1,
                          fact_marks_1, year_comment_1, comment_1, path):
        path_test = self.fill_GP_sample(2, '../tableTemplates/GP-GP.xlsx', indicator, event_id, event, main_event_id,
                                        main_event,
                                        subprogram_id, subprogram, type, unit, years, plan_marks, fact_marks,
                                        year_comment, comment, '../test1.xlsx')
        self.fill_GP_sample(6, path_test, indicator_1, event_id_1, event_1, main_event_id_1, main_event_1,
                            subprogram_id_1,
                            subprogram_1, type_1, unit_1, years_1, plan_marks_1, fact_marks_1, year_comment_1,
                            comment_1, path)
        os.remove(path_test)

    # def fill_CYR_CYR_sample(self, indicator_rf, task_id, task, cyr_id, cyr, indicator_vo, gosprogram, response_obj,
    #                         indicator_rf_1, task_id_1, task_1, cyr_id_1, cyr_1, indicator_vo_1, gosprogram_1, response_obj_1, path):
    #     path_test = self.fill_CYR_sample(2, '../tableTemplates/CYR-CYR.xlsx', indicator_rf, task_id, task, cyr_id,
    #                                      cyr, indicator_vo, gosprogram, response_obj,  '../test2.xlsx')
    #     self.fill_CYR_sample(5, path_test, indicator_rf_1, task_id_1, task_1, cyr_id_1, cyr_1, indicator_vo_1, gosprogram_1, response_obj_1, path)
    #     os.remove(path_test)

    """
    Все переменные должны быть типа str.
    indicator - наименование показателя,
    event_id - номер мероприятия, event - название мероприятия,
    main_event_id - номер основного мероприятия, main_event - название основного мероприятия,
    subprogram_id - номер подпрограммы, subprogram - название подпрограммы,
    type - тип (не поняла, что это именно - разберёшься),
    unit - единица измерения,
    years - список лет,
    plan_marks - список плановых значений показателя (порядок должен СОВПАДАТЬ с порядком лет),
    fact_marks - список фактических значений показателя (порядок должен СОВПАДАТЬ с порядком лет),
    year_comment - год комментария,
    comment - комментарий,
    path - путь к файлу
    """

    def fill_GP_sample(self, i, sample, indicator, event_id, event, main_event_id, main_event, subprogram_id,
                       subprogram,
                       type, unit, years, plan_marks, fact_marks, year_comment, comment, path):
        file = sample
        book = openpyxl.open(file)
        sheet = book.active
        sheet.cell(3, i + 1).value = indicator
        sheet.cell(4, i).value = 'Мероприятие №' + event_id if event_id is not None else 'Мероприятие'
        sheet.row_dimensions[4].height = 40
        sheet.cell(4, i + 1).value = event if event is not None else '-'
        sheet.cell(5,
                   i).value = 'Основное мероприятие №' + main_event_id if main_event_id is not None else 'Основное мероприятие'
        sheet.cell(5, i + 1).value = main_event if main_event is not None else '-'
        sheet.cell(6, i).value = 'Подпрограмма №' + subprogram_id
        sheet.cell(6, i + 1).value = subprogram
        sheet.cell(7, i + 1).value = type
        sheet.cell(8, i + 1).value = unit
        for j in range(len(years)):
            if j > 4:
                sheet.insert_rows(j + 11)
                sheet.row_dimensions[j + 11].height = 18
                font = Font(name='Times New Roman', size=12, bold=False, color='000000')
                medium = Side(border_style="medium", color="000000")
                thin = Side(border_style="thin", color="000000")
                name_style = NamedStyle(name="years")
                name_style.font = font
                name_style.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                sheet.cell(j + 11, i).border = Border(top=thin, bottom=thin, left=medium, right=thin)
                sheet.cell(j + 11, i + 1).border = Border(top=thin, bottom=thin, left=thin, right=thin)
                sheet.cell(j + 11, i + 2).border = Border(top=thin, bottom=thin, left=thin, right=medium)
                sheet.cell(j + 11, i).font = Font(name='Times New Roman', size=14, bold=False, color='000000')
                sheet.cell(j + 11, i + 1).font = font
                sheet.cell(j + 11, i + 2).font = font
                sheet.cell(j + 11, i).fill = PatternFill('solid', fgColor="e4efdc")
                sheet.cell(j + 11, i).alignment = Alignment(horizontal='center')
                sheet.cell(j + 11, i + 1).alignment = Alignment(horizontal='center')
                sheet.cell(j + 11, i + 2).alignment = Alignment(horizontal='center')
            sheet.cell(j + 11, i).value = years[j]
            print(j + 11, i + 1)
            sheet.cell(j + 11, i + 1).value = plan_marks[j]
            sheet.cell(j + 11, i + 2).value = fact_marks[j]
        sheet.cell(len(years) + 11, i).value = 'Комментарий (' + str(year_comment) + ' г.)'
        sheet.cell(len(years) + 11, i + 1).value = comment
        sheet.row_dimensions[len(years) + 11].height = 172.25
        create_diagram(sheet, 11, len(years) + 11, i, i + 2, i, len(years) + 16)
        book.save(path)
        return path

    """
    Все переменные должны быть типа str.
    indicator_rf - наименование показателя РФ, 
    task_id - номер задачи, 
    task - формулировка задачи, 
    cyr_id - номер ЦУРа, 
    cyr - формулировка ЦУР, 
    indicator_vo - наименование показателя ВО, 
    gosprogram - госпрограмма,
    response_obj - ответственный орган,
    path - путь к файлу

    """

    def fill_CYR_sample(self, i, sample, indicator_rf, task_id, task, cyr_id, cyr, indicator_vo, gosprogram,
                        response_obj, path):
        file = sample
        book = openpyxl.open(file)
        sheet = book.active
        sheet.cell(3, i + 1).value = indicator_rf
        sheet.cell(5, i).value = sheet.cell(5, i).value[:7] + '№' + task_id
        sheet.cell(5, i + 1).value = task
        sheet.cell(6, i).value = sheet.cell(6, i).value[:4] + '№' + cyr_id
        sheet.cell(6, i + 1).value = cyr
        sheet.cell(10, i + 1).value = indicator_vo
        sheet.cell(11, i + 1).value = gosprogram
        sheet.cell(12, i + 1).value = response_obj
        book.save(path)
        return path

# test = TableBlockConstructor('CYR', 'GP')
# # test.fill_CYR_sample("1", "1", "1", "1", "1", "1", "1", "1", "../test1.xlsx")
# test.fill_CYR_GP_sample('indicator', 'event_id', 'event', 'main_event_id', 'main_event', 'subprogram_id', 'subprogram',
#                         'type', 'unit', 'years', 'plan_marks', 'fact_marks', 'year_comment', 'comment',
#                         'indicator_rf', 'task_id', 'task', 'cyr_id', 'cyr', 'indicator_vo', 'gosprogram',
#                         'response_obj', '../CYR-GP.xlsx')
#
# years = [2011, 2012, 2013, 2014, 2015]
# years_1 = [2011, 2012, 2013, 2014, 2015, 2016, 2017]
# plan_marks = [20, 10, 45, 10, 20]
# fact_marks = [20, 30, 45, 70, 80]
#
# plan_marks_1 = [10, 100, 48, 10, 30, 5, 9]
# fact_marks_1 = [20, 20, 45, 56, 20, 0, 100]
#
# test.fill_GP_GP_sample('indicator', 'event_id', 'event', 'main_event_id', 'main_event', 'subprogram_id', 'subprogram',
#                        'type', 'unit', years, plan_marks, fact_marks, 'year_comment', 'comment', 'indicator_1',
#                        'event_id_1', 'event_1',
#                        'main_event_id_1', 'main_event_1', 'subprogram_id_1', 'subprogram_1', 'type_1', 'unit_1',
#                        years_1, plan_marks_1,
#                        fact_marks_1, 'year_comment_1', 'comment_1', '../GP-GP.xlsx')
# test.fill_CYR_CYR_sample('indicator_rf', 'task_id', 'task', 'cyr_id', 'cyr', 'indicator_vo', 'gosprogram', 'response_obj',
#                             'indicator_rf_1', 'task_id_1', 'task_1', 'cyr_id_1', 'cyr_1', 'indicator_vo_1', 'gosprogram_1',
#                          'response_obj_1', '../CYR-CYR.xlsx')
